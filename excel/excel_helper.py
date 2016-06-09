'''
Created on 31 May 2016

@author: Dev2
'''

from openpyxl import load_workbook
import warnings
import os

class ExcelHelper(object):
    '''
    Test Spread sheet reading code
    
    
    What I want to do is open the excel spread sheet check there is a value in Operation if not, stop the code
    
    read row=3
    read col=range[A:BJ]
    
    when test is complete:
        
        ##create an asset via the portal## mark as pass or fail and move onto the next row. break if row contains no operation
    
    read next row
    
    '''

    def __init__(self,excel_spreadsheet='C:\\workspace\\AtlasPortalAutomation\\excel\\excel_documents\\test.xlsm'):
        print('Reading Test Data from:'+excel_spreadsheet)
        self.excel_location=excel_spreadsheet
    
    '''
    @author: Dervis Suleyman
    @summary: reads a sheet from an excel spread sheet and returns the cell value
    '''
    def get_cell_value(self,sheet=None,row_number=None,col_number=None):
        cell_value=(sheet.cell(row=row_number, column=col_number).value)
        return cell_value
    
    '''
    @author: Derivs Suleyman
    @note: Handle the test set creation from spread sheet - need to identify if it's an update or delete test case
    '''
    def get_test_set(self):
        '''store each asset row as a test inside the test set'''
        test_set={"test_set" : {}}
        
        '''Ignore the warning for using .xlsm this is due to the python code expecting a 2006 extension'''
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            '''--------------------------------------------------------------------------------------------'''
            wb = load_workbook(filename=self.excel_location)
            print("-----------------------------------------")
            print("Active work sheet=["+str(wb.active)+"]")
            #read the Meta data work sheet
            sheet=wb.get_sheet_by_name('Metadata spreadsheet')
            #check the title of the work sheet can be found
            #print(sheet.title)
            print("-----------------------------------------")
            
            
            '''iterate over each row in the excel spread sheet need to handle utf-8 encoding'''
            for row_number in range(4, 100):
                '''stop the loop if there is no value in the operation field'''
                operation=self.get_cell_value(sheet, row_number, col_number=1)
                                              
                if (operation==None or len(operation)<=0):
                    print("EOF...stop reading the spread sheet")
                    print("-----------------------------------------")
                    return test_set

                
                '''Test Place Holder'''
                test='test'+str(row_number)
                '''update test set with new information'''
                test_set['test_set'].update({test:{}})
                
                '''Populate the series info based in information from the spread sheet'''
                series_info = {'SeriesID':self.get_cell_value(sheet, row_number, col_number=11),
                       'TitleBrief':self.get_cell_value(sheet, row_number, col_number=30),
                       'TitleMedium':self.get_cell_value(sheet, row_number, col_number=31),
                       'TitleLong':self.get_cell_value(sheet, row_number, col_number=32),
                       'SummaryBrief':self.get_cell_value(sheet, row_number, col_number=33),
                       'SummaryShort':self.get_cell_value(sheet, row_number, col_number=34),
                       'SummaryMedium':self.get_cell_value(sheet, row_number, col_number=35),
                       'SummaryLong':self.get_cell_value(sheet, row_number, col_number=36),
                       'Genre':self.get_cell_value(sheet, row_number, col_number=14),
                       'StudioDisplay':self.get_cell_value(sheet, row_number, col_number=15),
                       '16x9 image':'\\isilon\\test_images\\'+str(self.get_cell_value(sheet, row_number, col_number=58)),
                       '4x3 image':'\\isilon\\test_images\\'+str(self.get_cell_value(sheet, row_number, col_number=60))}
                
                test_set['test_set'][test].update({'series_info':series_info})
                
                season_info = {'SeasonID':self.get_cell_value(sheet, row_number, col_number=12),
                       'Series':self.get_cell_value(sheet, row_number, col_number=30),
                       'TitleBrief':self.get_cell_value(sheet, row_number, col_number=37),
                       'TitleMedium':self.get_cell_value(sheet, row_number, col_number=38),
                       'TitleLong':self.get_cell_value(sheet, row_number, col_number=39),
                       'SummaryBrief':self.get_cell_value(sheet, row_number, col_number=40),
                       'SummaryShort':self.get_cell_value(sheet, row_number, col_number=41),
                       'SummaryMedium':self.get_cell_value(sheet, row_number, col_number=42),
                       'SummaryLong':self.get_cell_value(sheet, row_number, col_number=43),
                       'SeasonNumber':self.get_cell_value(sheet, row_number, col_number=20),
                       'TotalEpisodes':self.get_cell_value(sheet, row_number, col_number=21),
                       'ProductionYear':str(self.get_cell_value(sheet, row_number, col_number=13)),
                       '16x9 image':'\\isilon\\test_images\\'+str(self.get_cell_value(sheet, row_number, col_number=59)),
                       '4x3 image':'\\isilon\\test_images\\'+str(self.get_cell_value(sheet, row_number, col_number=61))}
                
                test_set['test_set'][test].update({'season_info':season_info})
                
                portal_asset={'asset_id':self.get_cell_value(sheet, row_number, col_number=2),
                      'channel':self.get_cell_value(sheet, row_number, col_number=3),
                      'series':self.get_cell_value(sheet, row_number, col_number=30),
                      'season':self.get_cell_value(sheet, row_number, col_number=37),
                      'episodeNumber':self.get_cell_value(sheet, row_number, col_number=19),
                      'TotalEpisodes':self.get_cell_value(sheet, row_number, col_number=21),
                      'Title':self.get_cell_value(sheet, row_number, col_number=23),
                      'Summary':self.get_cell_value(sheet, row_number, col_number=26),
                      'Actors':self.get_cell_value(sheet, row_number, col_number=44),
                      'Warning':self.get_cell_value(sheet, row_number, col_number=17),
                      'DisplayRuntime':str(self.get_cell_value(sheet, row_number, col_number=18)),#may cause problems
                      'Genre':self.get_cell_value(sheet, row_number, col_number=14),
                      'Rating':self.get_cell_value(sheet, row_number, col_number=16),
                      'BroadcastDate':str(self.get_cell_value(sheet, row_number, col_number=5)),
                      'ProductionYear':str(self.get_cell_value(sheet, row_number, col_number=13)),
                      'Studio':self.get_cell_value(sheet, row_number, col_number=15),
                      'VideoFile':self.get_cell_value(sheet, row_number, col_number=46),
                      'SD/HD':self.get_cell_value(sheet, row_number, col_number=47),
                      'SOE':str(self.get_cell_value(sheet, row_number, col_number=48)),
                      'SOM':str(self.get_cell_value(sheet, row_number, col_number=49)),
                      'EOM':str(self.get_cell_value(sheet, row_number, col_number=50)),
                      'SOC':str(self.get_cell_value(sheet, row_number, col_number=51)),
                      'Duration':str(self.get_cell_value(sheet, row_number, col_number=52)),
                      'AspectRatio':self.get_cell_value(sheet, row_number, col_number=53),
                      'AudioTrackNumbers':str(self.get_cell_value(sheet, row_number, col_number=54)),
                      'SubTitles':self.get_cell_value(sheet, row_number, col_number=55),
                      '16-9-image':'\\isilon\\test_images\\'+str(self.get_cell_value(sheet, row_number, col_number=56)),
                      '4-3-image':'\\isilon\\test_images\\'+str(self.get_cell_value(sheet, row_number, col_number=57)),
                      'Boxart-image':'\\isilon\\test_images\\'+str(self.get_cell_value(sheet, row_number, col_number=62))}
                
                test_set['test_set'][test].update({'portal_asset':portal_asset})
                
                
                '''
                If offer from the spread sheet contains AT or AM or AS
                platform = "AT AM AS"
                platform = "AM AS"
                platform = "AT"
                '''
                
                platform = self.get_cell_value(sheet, row_number, col_number=7)
                
                if 'AT' in platform:
                    
                    offer={'Platform':'AT',
                    'Type':self.get_cell_value(sheet, row_number, col_number=8),
                    'StartDate':str(self.get_cell_value(sheet, row_number, col_number=5)),
                    'EndDate':str(self.get_cell_value(sheet, row_number, col_number=6))
                    }
                    
                    '''check for the platform key if it does not exist create it'''
                    if test_set['test_set'][test].has_key('platform')==False:
                        test_set['test_set'][test].update({'platform':{}})
                    
                    test_set['test_set'][test]['platform'].update({'AT':offer})
 
                
                if 'AM' in platform:
                    
                    offer={'Platform':'AM',
                    'Type':self.get_cell_value(sheet, row_number, col_number=8),
                    'StartDate':str(self.get_cell_value(sheet, row_number, col_number=5)),
                    'EndDate':str(self.get_cell_value(sheet, row_number, col_number=6))
                    }
                    
                    '''check for the platform key if it does not exist create it'''
                    if test_set['test_set'][test].has_key('platform')==False:
                        test_set['test_set'][test].update({'platform':{}})
                    
                    test_set['test_set'][test]['platform']['AM']=offer
                    
                if 'AS' in platform:
                    
                    offer={'Platform':'AS',
                    'Type':self.get_cell_value(sheet, row_number, col_number=8),
                    'StartDate':str(self.get_cell_value(sheet, row_number, col_number=5)),
                    'EndDate':str(self.get_cell_value(sheet, row_number, col_number=6))
                    }
                    
                    '''check for the platform key if it does not exist create it'''
                    if test_set['test_set'][test].has_key('platform')==False:
                        test_set['test_set'][test].update({'platform':{}})
                    
                    test_set['test_set'][test]['platform']['AS']=offer

        return test_set


if __name__ == '__main__':
    excel_helper = ExcelHelper('C:\\workspace\\AtlasPortalAutomation\\excel\\excel_documents\\test.xlsm')
    test_set = excel_helper.get_test_set()
    print(type(test_set))
    #print test_set['test_set']['test5']['series_info']['SummaryMedium']
    import json
    print json.dumps(test_set,indent=4, separators=(',', ': ')) 

'''
@note: for testing
'''
# if __name__ == '__main__':
#     with warnings.catch_warnings():
#         '''Ignore the warning for using .xlsm this is due to the python code expecting a 2006 extension'''
#         warnings.simplefilter("ignore")
#         '''--------------------------------------------------------------------------------------------'''
#         wb = load_workbook(filename='C:\\workspace\\AtlasPortalAutomation\\excel\\excel_documents\\test.xlsm')
#         print("-----------------------------------------")
#         print("Active work sheet=["+str(wb.active)+"]")
#         #read the Metadata work sheet
#         sheet=wb.get_sheet_by_name('Metadata spreadsheet')
#         #check the title of the work sheet can be found
#         print(sheet.title)
#         print("-----------------------------------------")
#         
#         '''iterate over each row'''
#         for row_number in range(4, 8):
#             for col_number in range(1,63):
#                 cell_value=(sheet.cell(row=row_number, column=col_number).value)
#                 if not cell_value==None:
#                     print cell_value
                    